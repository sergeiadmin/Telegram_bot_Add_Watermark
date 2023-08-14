import os
import telebot
from telebot import types
from PIL import Image
from openpyxl import Workbook, load_workbook
import logging
import datetime
import time
import signal
import sys

bot = telebot.TeleBot('6586120678:AAGZUQtr2zVw_fqigdhjfv0scrtkpcCD5oM')

# Настройка логирования
logging.basicConfig(level=logging.ERROR)
logger = logging.getLogger(__name__)


# Создание инлайн-кнопок на основе изображений в папке 'img'
def generate_buttons():
    buttons = []
    img_folder = 'img'
    if os.path.exists(img_folder) and os.path.isdir(img_folder):
        for img in os.listdir(img_folder):
            img_path = os.path.join(img_folder, img)
            if os.path.isfile(img_path):
                text = img.split('.')[0]
                callback = types.InlineKeyboardButton(text, callback_data=img_path)
                buttons.append(callback)
    markup = types.InlineKeyboardMarkup()
    markup.add(*buttons)
    return markup


# Обработчик ошибок
def handle_error(message, error_message):
    logger.error(error_message)
    bot.send_message(message.chat.id, error_message)


# Обработчик команды /start
@bot.message_handler(commands=['start'])
def start(message):
    markup = generate_buttons()
    bot.send_message(message.chat.id, 'Привет! Пришлите мне изображение или введите /help.', reply_markup=markup)


# Обработчик команды /help
@bot.message_handler(commands=['help'])
def help(message):
    bot.send_message(message.chat.id, '/start - Запустить бот\nПришлите мне фотографию, чтобы добавить водяной знак!')


# Обработчик изображений и документов
@bot.message_handler(content_types=['photo', 'document'])
def handle_image(message):
    try:
        # Обработка и сохранение изображения
        file_id = ''
        if message.content_type == 'photo':
            file_id = message.photo[-1].file_id
        elif message.content_type == 'document':
            if message.document.mime_type.startswith('image/'):
                file_id = message.document.file_id
            else:
                bot.send_message(message.chat.id, 'Пожалуйста, отправьте только изображение (JPG или PNG).')
                return
        file_info = bot.get_file(file_id)
        user_image_path = f'user_images/user_{message.from_user.id}.png'
        downloaded_file = bot.download_file(file_info.file_path)
        with open(user_image_path, 'wb') as new_file:
            new_file.write(downloaded_file)

        # Отправка инлайн-кнопок для выбора стиля
        markup = generate_buttons()
        bot.send_message(message.chat.id, 'Выберите стиль:', reply_markup=markup)
    except Exception as e:
        handle_error(message, 'Ошибка при обработке изображения.')


# Обработчик нажатий на инлайн-кнопки
@bot.callback_query_handler(func=lambda call: True)
def process_callback(call):
    img = call.data
    user_image_path = f'user_images/user_{call.from_user.id}.png'

    # Определение типа обработки на основе выбранной кнопки
    if os.path.basename(img) == "Студенческий билет.png":
        add_watermark(call, is_student_card=True)
    else:
        add_watermark(call, is_student_card=False)


# Добавление водяного знака или обработка студенческого билета
def add_watermark(call, is_student_card):
    try:
        img = call.data
        user_image_path = f'user_images/user_{call.from_user.id}.png'
        wm = Image.open(img)
        user_image = Image.open(user_image_path)

        if is_student_card:
            # Обработка студенческого билета с эффектом аналогичным object-fit: cover
            student_id_card = Image.open(img)

            target_width = 450
            target_height = 585

            # Вычисляем коэффициенты для масштабирования
            width_ratio = target_width / user_image.width
            height_ratio = target_height / user_image.height
            resize_ratio = max(width_ratio, height_ratio)

            # Масштабируем изображение пользователя
            new_width = int(user_image.width * resize_ratio)
            new_height = int(user_image.height * resize_ratio)
            user_image_resized = user_image.resize((new_width, new_height))

            # Определение координат для вставки с центрированием и обрезкой
            x_offset = 410
            y_offset = 560
            if new_width > target_width:
                x_offset -= (new_width - target_width) // 2
            if new_height > target_height:
                y_offset -= (new_height - target_height) // 2

            # Создаем новое изображение для вставки
            result = Image.new("RGB", student_id_card.size)
            result.paste(user_image_resized, (x_offset, y_offset))

            # Копируем содержимое водяного знака на новое изображение
            result.paste(wm, (0, 0), wm)

            result_filename = f'results/student_id_card_{call.from_user.id}.png'
            result.save(result_filename)
            with open(result_filename, 'rb') as photo:
                bot.send_photo(call.message.chat.id, photo)
            with open(result_filename, 'rb') as doc:
                bot.send_document(call.message.chat.id, doc)
        else:
            # Обработка водяного знака
            width_ratio = wm.width / user_image.width
            height_ratio = wm.height / user_image.height
            resize_ratio = max(width_ratio, height_ratio)
            new_width = int(user_image.width * resize_ratio)
            new_height = int(user_image.height * resize_ratio)
            img = user_image.resize((new_width, new_height))
            result = Image.new("RGB", wm.size)
            result.paste(img, ((wm.width - img.width) // 2, (wm.height - img.height) // 2))
            result.paste(wm, (0, 0), wm)
            result_filename = f'results/result_{call.from_user.id}.png'
            result.save(result_filename, format='jpeg', quality=90)
            with open(result_filename, 'rb') as photo:
                bot.send_photo(call.message.chat.id, photo)
                with open(result_filename, 'rb') as doc:
                    bot.send_document(call.message.chat.id, doc)

        append_to_excel(call.from_user)
    except Exception as e:
        handle_error(call.message, f'Ошибка при обработке изображения: {str(e)}')

# Добавление информации о пользователе в Excel
def append_to_excel(user):
    data_path = 'data'
    data_filename = f'{data_path}/data.xlsx'
    if not os.path.exists(data_path):
        os.makedirs(data_path)
    if not os.path.exists(data_filename):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "UserData"
        sheet.append(["Username", "User ID", "Date", "First_name", "Last_name"])
        workbook.save(data_filename)
    workbook = load_workbook(data_filename)
    if 'UserData' not in workbook.sheetnames:
        workbook.create_sheet(title='UserData')
        sheet = workbook['UserData']
        sheet.append(["Username", "User ID", "Date", "First_name", "Last_name"])
    else:
        sheet = workbook['UserData']
    sheet.append([
        user.username,
        user.id,
        str(datetime.datetime.now()),
        user.last_name,
        user.first_name,
    ])
    workbook.save(data_filename)


# Обработчик сигнала для корректного завершения программы
def signal_handler(sig, frame):
    print('Выход из программы')
    bot.stop_polling()
    sys.exit(0)


# Инициализация и запуск бота
def init():
    signal.signal(signal.SIGINT, signal_handler)

# Добавьте эту функцию для запуска бота
def run():
    while True:
        try:
            print("Бот запущен")
            bot.polling()
        except Exception as e:
            logger.error(f'Error in main loop: {e}')
            print("Перезапуск бота...")
            time.sleep(5)

if __name__ == '__main__':
    init()
    run()